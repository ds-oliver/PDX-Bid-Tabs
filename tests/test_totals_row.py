from openpyxl import Workbook

from bidtabs.parse_items import iter_item_rows
from bidtabs.parse_table import BidderBlock


def test_totals_row_detection():
    wb = Workbook()
    ws = wb.active

    ws.cell(4, 1, "Item No.")
    ws.cell(4, 2, "Item Description")
    ws.cell(4, 3, "Estimated Quantity")
    ws.cell(4, 4, "Units")
    ws.cell(4, 5, "Unit Price")
    ws.cell(4, 6, "Total Price")

    ws.cell(5, 1, "0010")
    ws.cell(5, 2, "Test Item")
    ws.cell(5, 3, 2)
    ws.cell(5, 4, "EA")
    ws.cell(5, 5, 3)
    ws.cell(5, 6, 6)

    ws.cell(6, 2, "Total Amount")
    ws.cell(6, 6, 123)

    col_map = {"line_no_col": 1, "desc_col": 2, "qty_col": 3, "unit_col": 4}
    bidder_blocks = [BidderBlock("ACME", 5, 6)]

    rows = list(iter_item_rows(ws, 4, col_map, bidder_blocks, termination_blank_streak=2))
    assert len(rows) == 2
    assert rows[0]["is_totals_row"] is False
    assert rows[1]["is_totals_row"] is True
    assert rows[1]["schedule_total"] == 123.0
