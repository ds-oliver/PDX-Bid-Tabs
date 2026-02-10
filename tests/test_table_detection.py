from openpyxl import Workbook

from bidtabs.parse_table import detect_table_header_row, map_base_columns


def test_detect_table_header_row():
    wb = Workbook()
    ws = wb.active
    ws.cell(5, 1, "Item No.")
    ws.cell(5, 2, "Item Description")
    ws.cell(5, 3, "Estimated Quantity")
    ws.cell(5, 4, "Units")
    ws.cell(5, 5, "Unit Price")
    ws.cell(5, 6, "Total Price")
    ws.cell(5, 7, "Unit Price")
    ws.cell(5, 8, "Total Price")

    row, status = detect_table_header_row(ws)
    assert status == "ok"
    assert row == 5

    cmap = map_base_columns(ws, row)
    assert cmap["line_no_col"] == 1
    assert cmap["desc_col"] == 2
    assert cmap["qty_col"] == 3
    assert cmap["unit_col"] == 4
