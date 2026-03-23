from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def test_pay_item_similarity_workbook_structure_and_unclassified():
    import subprocess
    import sys

    subprocess.check_call([
        sys.executable,
        "scripts/analyze_pay_item_similarity.py",
        "--input_csv",
        "./data_out/compiled_excel_itemized_clean.csv",
        "--output_csv",
        "./reports/pay_item_similarity_review.csv",
        "--output_xlsx",
        "./reports/pay_item_similarity_review.xlsx",
    ])

    xlsx = Path("reports/pay_item_similarity_review.xlsx")
    assert xlsx.exists()

    wb = load_workbook(xlsx, data_only=True)
    assert wb.sheetnames == ["README_Instructions", "Similarity_Summary", "HIGH_VARIANCE_DETAILS"]

    ws = wb["Similarity_Summary"]
    expected = [
        "Specification",
        "Spec Description",
        "Project Count",
        "Distinct Pay Item Descriptions",
        "Sample Pay Item Descriptions",
        "Distinct Supplemental Descriptions",
        "Sample Supplemental Descriptions",
        "Variance Level",
    ]
    got = [ws.cell(1, i).value for i in range(1, len(expected) + 1)]
    assert got == expected

    ws_high = wb["HIGH_VARIANCE_DETAILS"]
    expected_high = [
        "Item Description Raw",
        "Item",
        "Specification",
        "Spec Description",
        "Alternate Specification",
        "Pay Item Description",
        "Supplemental Description",
        "Unit",
        "Avg Unit Price",
        "Example Project EANs",
        "Matching Spec 1",
        "Matching Spec 2",
        "Matching Spec 3",
        "Reviewer Notes",
    ]
    got_high = [ws_high.cell(1, i).value for i in range(1, len(expected_high) + 1)]
    assert got_high == expected_high

    summary = pd.read_csv("./reports/pay_item_similarity_review.csv", keep_default_na=False)
    assert set(summary["Variance Level"].unique()).issubset({"HIGH", "LOW"})
    assert "UNCLASSIFIED" in set(summary["Specification"].astype(str))
    row = summary[summary["Specification"] == "UNCLASSIFIED"].iloc[0]
    assert row["Variance Level"] == "HIGH"

    row_012200 = summary[summary["Specification"] == "012200"]
    assert not row_012200.empty
    assert str(row_012200.iloc[0]["Spec Description"]).strip() != ""

    high_df = pd.read_excel("./reports/pay_item_similarity_review.xlsx", sheet_name="HIGH_VARIANCE_DETAILS")
    row_high_012200 = high_df[high_df["Specification"].astype(str) == "012200"]
    assert not row_high_012200.empty
    assert str(row_high_012200.iloc[0]["Spec Description"]).strip() != ""
