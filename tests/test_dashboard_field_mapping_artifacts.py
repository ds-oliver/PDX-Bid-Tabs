from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from bidtabs.reporting import DASHBOARD_FIELD_MAPPING_HEADERS as HEADERS


def test_dashboard_field_mapping_artifacts_exist_and_headers():
    import subprocess
    import sys

    subprocess.check_call([
        sys.executable,
        "scripts/build_dashboard_field_mapping.py",
        "--out_xlsx",
        "./docs/dashboard_field_mapping.xlsx",
    ])

    xlsx = Path("docs/dashboard_field_mapping.xlsx")
    assert xlsx.exists()

    wb = load_workbook(xlsx, data_only=True)
    assert wb.sheetnames == ["README", "Field_Mapping"]

    ws = wb["Field_Mapping"]
    got = [ws.cell(1, i).value for i in range(1, len(HEADERS) + 1)]
    assert got == HEADERS


def test_mapping_rows_use_bid_tab_business_sources():
    df = pd.read_excel("docs/dashboard_field_mapping.xlsx", sheet_name="Field_Mapping")

    src_text = " ".join(df["Port Data Object Source"].astype(str).tolist())
    for forbidden in ["Projects.", "Items.", "Bids.", "Projects.csv", "Items.csv", "Bids.csv"]:
        assert forbidden not in src_text

    expected_rows = {
        "District & County": "location_name_raw",
        "Specification": "specification",
        "Item": "item",
    }
    for odot, port_name in expected_rows.items():
        row = df[df["ODOT Report Object"] == odot]
        assert not row.empty
        assert row.iloc[0]["Port Data Object"] == port_name

    assert "Dev Notes" in df.columns
