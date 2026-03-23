from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "README",
    "EXEC_SUMMARY",
    "RAW_VARIABILITY_OVERVIEW",
    "RAW_SPEC_VARIABILITY",
    "QUALITY_ISSUES",
    "PROJECT_VARIABILITY",
    "SCHEDULE_COMPARISON",
    "RAW_VS_SMART_CLEAN",
    "TOP_ACTIONS",
]


def test_item_variability_report_artifacts_and_no_parse_confidence():
    import subprocess
    import sys

    # Ensure raw baseline source exists.
    subprocess.check_call([
        sys.executable,
        "scripts/build_raw_snapshot.py",
        "--input_dir",
        "./data_in",
        "--output_csv",
        "./data_out/compiled_excel_itemized_raw_snapshot.csv",
    ])

    subprocess.check_call([
        sys.executable,
        "scripts/build_item_variability_business_report.py",
        "--raw_snapshot_csv",
        "./data_out/compiled_excel_itemized_raw_snapshot.csv",
        "--clean_csv",
        "./data_out/compiled_excel_itemized_clean.csv",
        "--out_xlsx",
        "./reports/item_variability_business_report.xlsx",
        "--out_dir",
        "./reports",
        "--prefix",
        "item_variability",
    ])

    xlsx = Path("reports/item_variability_business_report.xlsx")
    assert xlsx.exists()

    wb = load_workbook(xlsx, data_only=True)
    assert wb.sheetnames == EXPECTED_SHEETS

    for csv_name in [
        "reports/item_variability_exec_summary.csv",
        "reports/item_variability_spec_summary.csv",
        "reports/item_variability_raw_vs_clean.csv",
        "reports/item_variability_quality_issues.csv",
        "reports/item_variability_rulebook.csv",
    ]:
        p = Path(csv_name)
        assert p.exists()
        df = pd.read_csv(p, keep_default_na=False)
        assert "parse_confidence" not in df.columns


def test_raw_baseline_metrics_unchanged_without_clean_source():
    import subprocess
    import sys

    subprocess.check_call([
        sys.executable,
        "scripts/build_item_variability_business_report.py",
        "--raw_snapshot_csv",
        "./data_out/compiled_excel_itemized_raw_snapshot.csv",
        "--clean_csv",
        "",
        "--out_xlsx",
        "./reports/item_variability_business_report_noclean.xlsx",
        "--out_dir",
        "./reports",
        "--prefix",
        "item_variability_noclean",
    ])

    with_clean = pd.read_csv("./reports/item_variability_spec_summary.csv", keep_default_na=False)
    no_clean = pd.read_csv("./reports/item_variability_noclean_spec_summary.csv", keep_default_na=False)

    cols = ["raw_spec_token", "row_count", "distinct_raw_desc_count", "raw_variability_ratio", "severity_band"]
    a = with_clean[cols].sort_values(cols, kind="stable").reset_index(drop=True)
    b = no_clean[cols].sort_values(cols, kind="stable").reset_index(drop=True)
    pd.testing.assert_frame_equal(a, b)
