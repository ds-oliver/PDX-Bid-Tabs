from openpyxl import Workbook

from bidtabs.parse_table import detect_bidder_blocks


def test_bidder_block_from_merged_headers():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=6)
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)
    ws.cell(3, 5, "ACME CONSTRUCTION")
    ws.cell(3, 7, "ENGINEER'S ESTIMATE")

    ws.cell(4, 5, "Unit Price")
    ws.cell(4, 6, "Total Price")
    ws.cell(4, 7, "Unit Price")
    ws.cell(4, 8, "Total Price")

    blocks, warnings = detect_bidder_blocks(ws, table_header_row=4)
    assert not warnings
    assert len(blocks) == 2
    assert blocks[0].bidder_name_raw == "ACME CONSTRUCTION"
    assert blocks[1].bidder_name_raw == "ENGINEER'S ESTIMATE"
