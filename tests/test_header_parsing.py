from openpyxl import Workbook

from bidtabs.parse_header import parse_project_header


def test_parse_project_header():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "PORTLAND INTERNATIONAL AIRPORT")
    ws.cell(2, 1, "TAXIWAY REHAB PROJECT")
    ws.cell(3, 1, "EAN 2023D016")
    ws.cell(3, 2, "Solicitation No. 12566")
    ws.cell(3, 3, "02/09/2026")

    h = parse_project_header(ws, table_header_row=10)
    assert h["project_ean"] == "2023D016"
    assert "12566" in h["solicitation_no"]
    assert h["letting_date"] == "2026-02-09"


def test_location_token_normalization():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "HILLSBORO AIRPORT")
    ws.cell(1, 6, "Name of Contractor 1")
    ws.cell(2, 1, "RUNWAY PROJECT")

    h = parse_project_header(ws, table_header_row=10)
    assert h["location_name_raw"] == "HILLSBOROAIRPORT"
