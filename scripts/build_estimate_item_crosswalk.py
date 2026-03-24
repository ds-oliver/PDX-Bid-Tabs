#!/usr/bin/env python3
"""Build estimate_item_crosswalk artifacts

Maps raw bid tab line item descriptions to standardized Port of Portland
Hard Cost Estimate Form items using spec-anchored fuzzy matching.

Run as: python build_estimate_item_crosswalk.py --input_csv ... --catalog_xlsx ... --out_dir ... --out_prefix ...
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from rapidfuzz import fuzz
    from rapidfuzz import process as rf_process
except ImportError:
    fuzz = None
    rf_process = None

# ================================================================
# PATHS  (defaults; may be overridden by CLI or library wrappers)
# ================================================================
CSV_PATH = Path("/Users/hogan/BTA 525 Capstone/PDX_BID_TABS_COMPLETE/data_out/compiled_excel_itemized_clean.csv")
CATALOG_XLSX = Path("/Users/hogan/BTA 525 Capstone/PDX_BID_TABS_COMPLETE/data_in/DRAFT One Port Estimating - Quantities Tool.xlsx")
OUTPUT_PATH = Path("/Users/hogan/BTA 525 Capstone/PDX_BID_TABS_COMPLETE/reports/estimate_item_crosswalk.xlsx")

# ================================================================
# SPEC BRIDGE: FAA spec → numeric code
# ================================================================
FAA_TO_NUMERIC: dict[str, int] = {
    "P-101": 510101, "P-152": 510152, "P-153": 510153, "P-154": 510154,
    "P-156": 510156, "P-209": 520209, "P-401": 540401, "P-403": 540403,
    "P-501": 550501, "P-602": 560602, "P-603": 560603, "P-604": 560604,
    "P-605": 560605, "P-606": 560606, "P-608": 560608, "P-609": 560609,
    "P-610": 560610, "P-620": 560620, "P-621": 560621, "P-626": 560626,
    "D-701": 580701, "D-705": 580705, "D-751": 580751, "D-752": 580752,
    "F-162": 570162,
    "T-901": 590901, "T-905": 590905, "T-908": 590908,
    "L-108": 600108, "L-110": 600110,
}

# ================================================================
# OUT-OF-CATALOG-SCOPE: spec → slicer_family
# These specs have no Hard Cost Estimate Form entries.
# ================================================================
OUT_OF_SCOPE: dict[str, str] = {
    "344300": "airfield_lighting",
    "344301": "airfield_lighting",
    "344302": "airfield_lighting",
    "260500": "electrical",
    "260526": "electrical",
    "260543": "electrical",
    "262200": "electrical",
    "262400": "electrical",
    "265000": "electrical",
    "265623": "electrical",
    "260553": "electrical",
    "263343": "ev_charging",
    "271000": "telecom_data",
    "270553": "telecom_data",
    "282300": "security_systems",
    "347113": "site_furnishings",
    "347115": "site_furnishings",
    "111100": "cable_retrievers",
    "354000": "waterway",
    "328400": "irrigation",
    "329300": "planting",
    "600108": "cable_electrical_faa",
    "600110": "cable_electrical_faa",
    "L-108": "cable_electrical_faa",
    "L-110": "cable_electrical_faa",
    "C-102": "other_faa",
    "C-105": "other_faa",
}

# Override C GR keyword → catalog item_description
GR_OVERRIDE_MAP: dict[str, str] = {
    "escort": "Escort",
    "security control": "Security Control Monitoring",
    "traffic control": "Work Zone and Traffic Control",
    "sweeping": "Sweeping",
    "construction survey": "Construction Survey",
    "mobilization": "Mobilization & Demobilization",
    "demobilization": "Mobilization & Demobilization",
}

# Detail sheet column order
DETAIL_COLUMNS = [
    "project_ean", "line_no", "item_description_raw", "specification",
    "pay_item_description", "unit_code_norm",
    "estimate_item", "division", "cost_code", "cost_code_desc",
    "needs_slicer", "slicer_family", "size_in", "material_or_type", "spec_family",
    "match_method", "match_confidence", "mapping_status", "needs_review",
]


# ================================================================
# TEXT HELPERS
# ================================================================

def _clean(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _norm(v: object) -> str:
    """Normalise to upper-case word tokens for fuzzy comparison."""
    t = str(v or "")
    t = re.sub(r"([a-z])([A-Z])", r"\1 \2", t)          # camelCase split
    t = re.sub(r"([A-Za-z])(\d)", r"\1 \2", t)
    t = re.sub(r"(\d)([A-Za-z])", r"\1 \2", t)
    t = re.sub(r"[^A-Za-z0-9]+", " ", t)
    return re.sub(r"\s+", " ", t).strip().upper()


def _lookup_key(v: object) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(v or "")).upper()


def _norm_numeric(v: object) -> str:
    s = str(v or "").strip()
    if re.fullmatch(r"\d+(\.0+)?", s):
        s = s.split(".")[0]
    if re.fullmatch(r"\d{1,6}", s):
        return s.zfill(6)
    return s


def _is_faa_spec(spec: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]-\d{3}[a-z]*", str(spec).strip().upper()))


def _score(a: str, b: str) -> float:
    if fuzz is not None:
        return float(fuzz.token_set_ratio(a, b))
    sa, sb = set(a.split()), set(b.split())
    if not sa or not sb:
        return 0.0
    return round(100.0 * len(sa & sb) / len(sa | sb), 2)


# ================================================================
# SLICER EXTRACTION
# ================================================================

def _extract_size_in(raw: str) -> str:
    pats = [
        r'(\d+(?:\.\d+)?)\s*(?:-\s*)?inch(?:es)?(?:\s*thick|\s*depth|\s*maximum\s*depth)?',
        r'(\d+(?:\.\d+)?)"',
        r"(\d+(?:\.\d+)?)'",
    ]
    for p in pats:
        m = re.search(p, raw, re.IGNORECASE)
        if m:
            return m.group(1)
    return ""


def _extract_material(raw: str) -> str:
    ordered = [
        ("non-perforated", "NON-PERFORATED"),
        ("non perforated", "NON-PERFORATED"),
        ("perforated", "PERFORATED"),
        ("\\bsolid\\b", "SOLID"),
        ("\\bpcc\\b", "PCC"),
        ("portland cement concrete", "PCC"),
        ("\\bhmac\\b", "HMAC"),
        ("\\bhma\\b", "HMA"),
        ("\\bhdpe\\b", "HDPE"),
        ("\\brcp\\b", "RCP"),
        ("\\bsts\\b", "STS"),
        ("\\bcmp\\b", "CMP"),
        ("\\bdip\\b", "DIP"),
        ("\\bac\\b", "AC"),
        ("bituminous", "BITUMINOUS"),
        ("asphalt", "ASPHALT"),
        ("concrete", "CONCRETE"),
    ]
    r = str(raw or "")
    for pat, label in ordered:
        if re.search(pat, r, re.IGNORECASE):
            return label
    return ""


# ================================================================
# CATALOG LOADING
# ================================================================

def load_catalog() -> tuple[pd.DataFrame, dict[str, str]]:
    """
    Returns:
        catalog_df  – one row per unique (cost_code, item_description)
        spec_to_numeric – maps spec strings to 6-digit numeric code strings
    """
    import openpyxl
    wb = openpyxl.load_workbook(CATALOG_XLSX, data_only=True)

    # --- Build desc→numeric lookup from DivisionCostCodeLookup (cols F, G) ---
    ws_lu = wb["DivisionCostCodeLookup"]
    desc_to_numeric: dict[str, str] = {}
    spec_to_numeric: dict[str, str] = {k: str(v).zfill(6) for k, v in FAA_TO_NUMERIC.items()}

    for row in ws_lu.iter_rows(min_row=2, values_only=True):
        if len(row) < 7:
            continue
        f_val, g_val = row[5], row[6]
        if f_val is None or g_val is None:
            continue
        ncode = _norm_numeric(g_val)
        if not ncode or not ncode.isdigit():
            continue
        key = _lookup_key(f_val)
        if key:
            desc_to_numeric[key] = ncode
        # Also register any FAA or numeric spec embedded in parenthetical
        faa_m = re.search(r"\(([A-Z]-\d{3}[a-z]*)\)", str(f_val), re.IGNORECASE)
        if faa_m:
            spec_to_numeric[faa_m.group(1).upper()] = ncode
        num_m = re.search(r"\((\d{4,6})\)", str(f_val))
        if num_m:
            spec_to_numeric[num_m.group(1).zfill(6)] = ncode

    # --- Parse Hard Cost Estimate Form (cols A, B, D) ---
    ws_hc = wb["Hard Cost Estimate Form"]

    SKIP_D_PATTERNS = re.compile(
        r"^(P|D|T|F|L|C)-\d{3}|^XX-|^Division\s+|^Front\s+End\s+By",
        re.IGNORECASE,
    )

    records: list[dict] = []
    seen_dedup: set[tuple] = set()

    for row in ws_hc.iter_rows(min_row=8, values_only=True):
        a = row[0] if len(row) > 0 else None  # Division
        b = row[1] if len(row) > 1 else None  # Cost Code Description
        d = row[3] if len(row) > 3 else None  # Item Description

        if not a or not d:
            continue
        division = _clean(a)
        if not division or not division[0].isdigit():
            continue

        item_desc = _clean(d)
        if not item_desc:
            continue
        if SKIP_D_PATTERNS.search(item_desc):
            continue
        # Skip pure spec labels ("P-101", "D-701", etc.)
        if re.fullmatch(r"[PDTFLCM]-\d{3}[a-z]*", item_desc.strip(), re.IGNORECASE):
            continue
        # Strip "(Section N)" suffixes
        item_clean = re.sub(r"\(section\s*\d+\)", "", item_desc, flags=re.IGNORECASE).strip()
        item_clean = re.sub(r"\s+", " ", item_clean).strip()

        cost_code_desc = _clean(b) if b else ""

        # Resolve numeric_code from cost_code_desc
        dkey = _lookup_key(cost_code_desc)
        numeric_code = desc_to_numeric.get(dkey, "")

        if not numeric_code:
            faa_m = re.search(r"\(([A-Z]-\d{3}[a-z]*)\)", cost_code_desc, re.IGNORECASE)
            if faa_m:
                fs = faa_m.group(1).upper()
                numeric_code = spec_to_numeric.get(fs, FAA_TO_NUMERIC.get(fs, ""))
                if numeric_code:
                    numeric_code = str(numeric_code).zfill(6)

        if not numeric_code:
            num_m = re.search(r"\b(\d{4,6})\b", cost_code_desc)
            if num_m:
                numeric_code = num_m.group(1).zfill(6)

        if not numeric_code:
            numeric_code = ""

        is_faa = "FAA" in division.upper() or bool(
            re.search(r"\([A-Z]-\d{3}", cost_code_desc, re.IGNORECASE)
        )

        # Build match_target = cleaned cost_code_desc + item_description
        ccd_clean = re.sub(r"\([^)]+\)", "", cost_code_desc).strip()  # strip parenthetical
        match_target = _norm(f"{ccd_clean} {item_clean}")

        # estimate_item
        estimate_item = f"{numeric_code} {item_clean}".strip() if numeric_code else item_clean

        # Dedup on (numeric_code, item_clean)
        key = (numeric_code, item_clean)
        if key in seen_dedup:
            continue
        seen_dedup.add(key)

        # needs_slicer: items with placeholder dimension (0'' or XX-)
        needs_slicer = bool(
            re.search(r"0''\s*(Portland|Cement|PCC)", item_desc, re.IGNORECASE)
            or "XX-" in item_desc
        )

        records.append({
            "estimate_item": estimate_item,
            "division": division,
            "cost_code": numeric_code,
            "cost_code_desc": cost_code_desc,
            "item_description": item_clean,
            "is_faa": is_faa,
            "match_target": match_target,
            "needs_slicer": needs_slicer,
        })

    cat = pd.DataFrame(records)
    if cat.empty:
        raise ValueError("Catalog extraction produced no rows — check Hard Cost Estimate Form.")
    return cat, spec_to_numeric


# ================================================================
# BID TAB LOADING
# ================================================================

def load_bid_tabs() -> pd.DataFrame:
    use_cols = [
        "project_ean", "line_no", "item_description_raw", "specification",
        "alternate_specification", "pay_item_description", "unit_code_norm",
        "unit_code_raw", "is_totals_row",
    ]
    raw = pd.read_csv(CSV_PATH, keep_default_na=False,
                      usecols=lambda c: c in use_cols)

    # Exclude totals rows
    totals = raw.get("is_totals_row", pd.Series(False)).astype(str).str.lower().isin(["true", "1"])
    df = raw[~totals].copy()

    # Deduplicate to unique (project_ean, line_no)
    df["project_ean"] = df["project_ean"].astype(str).str.strip()
    df["line_no"] = df["line_no"].astype(str).str.strip()
    df = df.drop_duplicates(subset=["project_ean", "line_no"], keep="first").reset_index(drop=True)

    df["item_description_raw"] = df["item_description_raw"].astype(str).map(_clean)
    df["specification"] = df["specification"].astype(str).str.strip().str.upper()
    df["pay_item_description"] = df["pay_item_description"].astype(str).map(_clean)

    unit = df.get("unit_code_norm", pd.Series("")).astype(str).str.strip()
    unit_raw = df.get("unit_code_raw", pd.Series("")).astype(str).str.strip()
    unit = unit.where(unit.str.strip() != "", unit_raw)
    df["unit_code_norm"] = unit

    # Build match query: pay_item_description (normalised), fallback to raw
    query = df["pay_item_description"].where(
        df["pay_item_description"].str.strip() != "",
        df["item_description_raw"]
    ).map(_norm)
    df["_query"] = query

    # Parse all spec tokens (primary + alternate + raw parentheticals)
    df["_specs"] = df.apply(
        lambda r: _parse_specs(
            r["item_description_raw"],
            r["specification"],
            r.get("alternate_specification", ""),
        ),
        axis=1,
    )
    return df


def _parse_specs(raw_desc: str, primary_spec: str, alt_spec: str) -> list[str]:
    specs: list[str] = []
    seen: set[str] = set()

    def add(s: str) -> None:
        s = s.strip().upper()
        if s and s not in seen:
            specs.append(s)
            seen.add(s)

    # Primary
    p = primary_spec.strip().upper()
    if p and p != "UNCLASSIFIED":
        add(p)

    # Alternate (may contain pipes or spaces)
    for part in re.split(r"[\|,;&\s]+", str(alt_spec or "")):
        part = part.strip().upper()
        if re.fullmatch(r"[A-Z]-\d{3}[A-Z]?", part) or re.fullmatch(r"\d{4,6}", part):
            add(part.zfill(6) if part.isdigit() else part)

    # Scan raw description for FAA specs and 6-digit section codes
    for m in re.finditer(r"\b([A-Z]-\d{3}[a-z]*)\b", raw_desc, re.IGNORECASE):
        add(m.group(1).upper())
    for m in re.finditer(r"\b(\d{6})\b", raw_desc):
        add(m.group(1))
    # "Section(s) XXXXXX" patterns
    for m in re.finditer(r"sections?\s+([\d\s,and]+)", raw_desc, re.IGNORECASE):
        for n in re.findall(r"\d{4,6}", m.group(1)):
            add(n.zfill(6))

    # If still empty, mark UNCLASSIFIED
    if not specs:
        specs = ["UNCLASSIFIED"]
    return specs


# ================================================================
# SPEC → NUMERIC RESOLUTION
# ================================================================

def _spec_to_numeric_code(spec: str, bridge: dict[str, str]) -> str | None:
    """Return numeric code string or None if unresolvable/out-of-scope."""
    s = spec.strip().upper()
    # Out-of-scope specs → None (caller handles oos_family)
    if s in OUT_OF_SCOPE:
        return None
    # FAA bridge (also handles L-108/L-110 which are also in OUT_OF_SCOPE above)
    if s in bridge:
        return bridge[s]
    # 6-digit numeric section codes → themselves
    if re.fullmatch(r"\d{6}", s):
        return s
    # 4-5 digit: zero-pad to 6
    if re.fullmatch(r"\d{4,5}", s):
        return s.zfill(6)
    return None


def _candidate_pool(
    specs: list[str],
    cat: pd.DataFrame,
    bridge: dict[str, str],
) -> tuple[pd.DataFrame | None, str | None]:
    """
    Returns (pool_df, oos_family).
    - If all specs out-of-scope → (None, family)
    - If some specs resolve to empty catalog pools → out_of_catalog_scope
    - If UNCLASSIFIED → (full catalog, None)
    """
    is_unclassified = (specs == ["UNCLASSIFIED"])
    if is_unclassified:
        return cat, None

    pools: list[pd.DataFrame] = []
    oos_families: list[str] = []
    any_resolvable = False

    for spec in specs:
        s = spec.strip().upper()
        if s == "UNCLASSIFIED":
            return cat, None  # open match
        if s in OUT_OF_SCOPE:
            oos_families.append(OUT_OF_SCOPE[s])
            continue
        ncode = _spec_to_numeric_code(s, bridge)
        if ncode is None:
            oos_families.append("out_of_catalog_scope")
            continue
        pool = cat[cat["cost_code"] == ncode]
        any_resolvable = True
        if not pool.empty:
            pools.append(pool)
        else:
            # Spec resolves to a code but no catalog entries → out of scope
            oos_families.append("out_of_catalog_scope")

    if pools:
        combined = pd.concat(pools, ignore_index=True).drop_duplicates(
            subset=["estimate_item"], keep="first"
        )
        return combined, None

    if not any_resolvable and oos_families:
        return None, oos_families[0]

    return None, oos_families[0] if oos_families else "out_of_catalog_scope"


# ================================================================
# MATCHING
# ================================================================

def _faa_signal(specs: list[str]) -> bool:
    return any(_is_faa_spec(s) for s in specs)


def _best_match(
    query: str,
    pool: pd.DataFrame,
) -> tuple[float, dict]:
    if rf_process is not None and fuzz is not None:
        idx_to_row = {i: r.to_dict() for i, (_, r) in enumerate(pool.iterrows())}
        choices = {i: v["match_target"] for i, v in idx_to_row.items()}
        extracted = rf_process.extract(
            query, choices, scorer=fuzz.token_set_ratio, limit=min(20, len(choices))
        )
        cands = [(float(sc), idx_to_row[idx]) for _, sc, idx in extracted]
    else:
        cands = [(
            _score(query, r["match_target"]),
            r.to_dict(),
        ) for _, r in pool.iterrows()]

    cands.sort(key=lambda x: -x[0])
    return cands[0] if cands else (0.0, {})


def _tie_break(
    cands: list[tuple[float, dict]],
    faa_sig: bool,
    non_faa_sig: bool,
) -> tuple[float, dict]:
    if not cands:
        return 0.0, {}
    cands.sort(key=lambda x: -x[0])
    top_score = cands[0][0]
    tied = [c for c in cands if top_score - c[0] <= 2]

    if faa_sig and not non_faa_sig:
        pref = [t for t in tied if t[1].get("is_faa", False)]
        if pref:
            tied = pref
    elif non_faa_sig and not faa_sig:
        pref = [t for t in tied if not t[1].get("is_faa", False)]
        if pref:
            tied = pref

    tied.sort(key=lambda x: len(str(x[1].get("item_description", ""))))
    return tied[0]


def match_row(
    row: pd.Series,
    cat: pd.DataFrame,
    bridge: dict[str, str],
) -> dict:
    """Apply the full 4-step matching algorithm to one bid tab row."""
    specs: list[str] = row["_specs"]
    query: str = row["_query"]
    primary_spec = row["specification"].strip().upper()
    raw_desc = row["item_description_raw"]
    pay_lower = _norm(row["pay_item_description"]).lower()

    faa_sig = _faa_signal(specs)
    non_faa_sig = not faa_sig and primary_spec not in ("UNCLASSIFIED",) and not _is_faa_spec(primary_spec)

    # ----------------------------------------------------------------
    # HARDCODED OVERRIDES  (applied before fuzzy matching)
    # ----------------------------------------------------------------

    # Override A — PCC removal (FAA vs Non-FAA)
    is_pcc = "PCC" in _norm(raw_desc) or "PORTLAND CEMENT CONCRETE" in _norm(raw_desc).replace("-", " ")
    is_removal = "REMOVAL" in _norm(raw_desc)

    if is_pcc and is_removal:
        if primary_spec == "P-101":
            m = cat[(cat["cost_code"] == "510101") &
                    cat["item_description"].str.upper().str.contains("PORTLAND CEMENT|PCC", na=False, regex=True)]
            if not m.empty:
                return _hit(m.iloc[0], 100, "hardcoded_override", False, raw_desc)
        elif primary_spec in ("024113", "UNCLASSIFIED"):
            m = cat[(cat["cost_code"] == "024113") &
                    cat["item_description"].str.upper().str.contains("PORTLAND CEMENT|PCC", na=False, regex=True)]
            if not m.empty:
                return _hit(m.iloc[0], 100, "hardcoded_override", False, raw_desc)

    # Override B — Asphalt removal under P-101
    if primary_spec == "P-101":
        if "ASPHALT" in _norm(raw_desc) and is_removal and not is_pcc:
            m = cat[(cat["cost_code"] == "510101") &
                    cat["item_description"].str.upper().str.contains("MILLING", na=False)]
            if not m.empty:
                return _hit(m.iloc[0], 90, "hardcoded_override", True, raw_desc)

    # Override C — Standard GR items under 012200
    if primary_spec == "012200":
        for kw, target_desc in GR_OVERRIDE_MAP.items():
            if kw in pay_lower or kw in _norm(raw_desc).lower():
                m = cat[cat["item_description"].str.strip().str.lower() == target_desc.lower()]
                if not m.empty:
                    return _hit(m.iloc[0], 95, "hardcoded_override", False, raw_desc)

    # Override D — Structure adjustment under 312300
    if primary_spec == "312300":
        if re.search(r"\b(adjust|adjustment|grade)\b", raw_desc, re.IGNORECASE):
            m = cat[(cat["cost_code"] == "312300") &
                    cat["item_description"].str.upper().str.contains("STABILIZATION", na=False)]
            if not m.empty:
                return _hit(m.iloc[0], 85, "hardcoded_override", True, raw_desc)

    # ----------------------------------------------------------------
    # STEP 1 — Resolve spec to candidate pool
    # ----------------------------------------------------------------
    pool, oos_family = _candidate_pool(specs, cat, bridge)

    if pool is None:
        return {
            "estimate_item": "", "division": "", "cost_code": "",
            "cost_code_desc": "", "needs_slicer": False,
            "slicer_family": oos_family or "out_of_catalog_scope",
            "size_in": "", "material_or_type": "", "spec_family": "",
            "match_method": "out_of_catalog_scope",
            "match_confidence": 0.0,
            "mapping_status": "out_of_catalog_scope",
            "needs_review": False,
        }

    if pool.empty:
        return {
            "estimate_item": "", "division": "", "cost_code": "",
            "cost_code_desc": "", "needs_slicer": False,
            "slicer_family": "out_of_catalog_scope",
            "size_in": "", "material_or_type": "", "spec_family": "",
            "match_method": "out_of_catalog_scope",
            "match_confidence": 0.0,
            "mapping_status": "out_of_catalog_scope",
            "needs_review": False,
        }

    is_unclassified = (specs == ["UNCLASSIFIED"])

    # ----------------------------------------------------------------
    # STEP 2 — Fuzzy match within pool
    # ----------------------------------------------------------------
    if rf_process is not None and fuzz is not None:
        idx_map = {i: r.to_dict() for i, (_, r) in enumerate(pool.iterrows())}
        choices = {i: v["match_target"] for i, v in idx_map.items()}
        extracted = rf_process.extract(
            query, choices, scorer=fuzz.token_set_ratio, limit=min(20, len(choices))
        )
        cands: list[tuple[float, dict]] = [(float(sc), idx_map[idx]) for _, sc, idx in extracted]
    else:
        cands = [(_score(query, r["match_target"]), r.to_dict()) for _, r in pool.iterrows()]

    if not cands:
        return _unmapped(0.0)

    # ----------------------------------------------------------------
    # STEP 4 (tie-breaking)
    # ----------------------------------------------------------------
    best_score, best = _tie_break(cands, faa_sig, non_faa_sig)
    if not best:
        return _unmapped(0.0)

    # ----------------------------------------------------------------
    # STEP 3 — Assign based on score and pool type
    # ----------------------------------------------------------------
    if is_unclassified:
        if best_score >= 88:
            status, review = "mapped", False
            method = "fuzzy_open"
        elif best_score >= 70:
            status, review = "mapped", True
            method = "fuzzy_open"
        else:
            return _unmapped(best_score)
    else:
        # Spec-anchored: always assign
        status = "mapped"
        review = best_score < 75
        method = "fuzzy_spec_anchored"

    return _hit(best, best_score, method, review, raw_desc, status)


def _hit(cat_row: dict, score: float, method: str, needs_review: bool,
         raw_desc: str, status: str = "mapped") -> dict:
    size_in = material = spec_fam = ""
    if cat_row.get("needs_slicer"):
        size_in = _extract_size_in(raw_desc)
        material = _extract_material(raw_desc)
        spec_fam = "FAA" if cat_row.get("is_faa") else "NON_FAA"
    elif cat_row.get("estimate_item", ""):
        spec_fam = "FAA" if cat_row.get("is_faa") else "NON_FAA"
    return {
        "estimate_item": str(cat_row.get("estimate_item", "")),
        "division": str(cat_row.get("division", "")),
        "cost_code": str(cat_row.get("cost_code", "")),
        "cost_code_desc": str(cat_row.get("cost_code_desc", "")),
        "needs_slicer": bool(cat_row.get("needs_slicer", False)),
        "slicer_family": "",
        "size_in": size_in,
        "material_or_type": material,
        "spec_family": spec_fam,
        "match_method": method,
        "match_confidence": round(float(score), 2),
        "mapping_status": status,
        "needs_review": needs_review,
    }


def _unmapped(score: float) -> dict:
    return {
        "estimate_item": "", "division": "", "cost_code": "",
        "cost_code_desc": "", "needs_slicer": False,
        "slicer_family": "", "size_in": "", "material_or_type": "",
        "spec_family": "",
        "match_method": "unmapped",
        "match_confidence": round(float(score), 2),
        "mapping_status": "unmapped",
        "needs_review": True,
    }


# ================================================================
# CORE PIPELINE
# ================================================================

def run_crosswalk() -> tuple[pd.DataFrame, pd.DataFrame]:
    cat, bridge = load_catalog()
    items = load_bid_tabs()

    results = [match_row(row, cat, bridge) for _, row in items.iterrows()]

    detail = items.copy()
    for col in ["estimate_item", "division", "cost_code", "cost_code_desc",
                "needs_slicer", "slicer_family", "size_in", "material_or_type",
                "spec_family", "match_method", "match_confidence",
                "mapping_status", "needs_review"]:
        detail[col] = [r[col] for r in results]

    # Drop internal helper columns
    for c in ["_query", "_specs"]:
        if c in detail.columns:
            detail.drop(columns=[c], inplace=True)

    detail = detail[DETAIL_COLUMNS].sort_values(
        ["project_ean", "line_no"], kind="stable"
    ).reset_index(drop=True)

    # ----------------------------------------------------------------
    # Crosswalk_Table: one row per (estimate_item, item_description_raw)
    # ----------------------------------------------------------------
    mapped = detail[detail["estimate_item"].str.strip() != ""].copy()
    mapped["_pk"] = mapped["project_ean"] + "|" + mapped["line_no"]
    grp = (
        mapped.groupby(["estimate_item", "item_description_raw"], dropna=False)
        .agg(
            bid_tab_count=("_pk", "nunique"),
            project_ean=("project_ean", lambda s: "; ".join(
                sorted({str(v).strip() for v in s if str(v).strip()})
            )),
        )
        .reset_index()
    )
    table = grp.rename(columns={
        "estimate_item": "Estimate Item",
        "bid_tab_count": "Bid Tab Count",
        "project_ean": "Project EAN",
        "item_description_raw": "Raw Bid Tab Item Description",
    })[["Estimate Item", "Bid Tab Count", "Project EAN", "Raw Bid Tab Item Description"]]
    table = table.sort_values(
        ["Estimate Item", "Bid Tab Count"],
        ascending=[True, False],
        kind="stable",
    ).reset_index(drop=True)

    return detail, table


# ================================================================
# EXCEL OUTPUT
# ================================================================

README_TEXT = """ESTIMATE ITEM CROSSWALK — READER'S GUIDE
=========================================

PURPOSE
-------
This workbook maps every raw line item from Port of Portland bid tabs to a
standardized item in the Hard Cost Estimate Form catalog. It helps answer:
"How many times has each catalog item appeared in real bids, and in which projects?"

SHEETS
------
Sheet 2 – Crosswalk_Table (PRIMARY DELIVERABLE)
  One row per unique combination of Estimate Item and raw bid description.
  Columns:
    Estimate Item             The standardized catalog label (e.g. "13501 Escort").
    Bid Tab Count             Number of unique project/line combinations that matched.
    Project EAN               Semicolon-separated list of contributing project EANs.
    Raw Bid Tab Item Description  The exact original text from the bid tab.

Sheet 3 – Detail
  One row per deduplicated bid tab line item (project_ean + line_no).
  Shows every matching decision: estimate_item, division, cost_code, match method,
  confidence score, mapping status, and slicer data.

Sheet 4 – Unmapped
  Items where no sufficiently confident match could be found (UNCLASSIFIED spec
  below score threshold). These warrant manual review or catalog expansion.

Sheet 5 – Out_of_Catalog_Scope
  Items whose specification code is valid but has no entry in the Hard Cost
  Estimate Form (e.g. airfield lighting, electrical, telecom). These are NOT
  matching failures — they represent work categories outside the current catalog.

Sheet 6 – Slicer_QA
  Rows where the matched catalog item has a variable dimension (pipe size,
  pavement thickness). Shows extracted size, material type, and FAA/Non-FAA flag
  for quality review.

UNDERSTANDING MAPPING STATUS
-----------------------------
  mapped               — Matched to a catalog item (review Bid Tab Count for confidence).
  unmapped             — Could not match (UNCLASSIFIED + low fuzzy score). Needs manual review.
  out_of_catalog_scope — Spec is valid but the catalog has no matching division/item.

UNDERSTANDING NEEDS_REVIEW
---------------------------
  True  — Match confidence was below the threshold (< 75 for spec-anchored,
          < 88 for open/UNCLASSIFIED). A human should verify the assignment.
  False — High-confidence match; no review needed.

BID TAB COUNT SEMANTICS
------------------------
  Each row is counted as a unique (project_ean, line_no) pair.
  Multiple bidders on the same line item count as ONE because the line item
  itself — not the bid price — is what we're cataloging.

DATA SCOPE
----------
  Source bid tabs: compiled_excel_itemized_clean.csv
  Catalog: DRAFT One Port Estimating - Quantities Tool.xlsx (Hard Cost Estimate Form)
  Matching engine: rapidfuzz token_set_ratio with spec-anchored candidate pools
"""


def _hdr(ws, row: int = 1) -> None:
    fill = PatternFill(fill_type="solid", start_color="FF1F4E79", end_color="FF1F4E79")
    font = Font(color="FFFFFFFF", bold=True)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font


def _autowidth(ws, max_w: int = 80) -> None:
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, max_w)


def _df_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    ws.freeze_panes = "A2"
    _hdr(ws)
    _autowidth(ws)


def write_xlsx(detail: pd.DataFrame, table: pd.DataFrame) -> None:
    unmapped = detail[detail["mapping_status"] == "unmapped"].copy()
    oos = detail[detail["mapping_status"] == "out_of_catalog_scope"].copy()
    slicer_qa = detail[
        detail["needs_slicer"].astype(bool) & (detail["mapping_status"] == "mapped")
    ][["estimate_item", "slicer_family", "item_description_raw",
       "size_in", "material_or_type", "spec_family", "needs_review"]].copy()

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: README ----
    readme = wb.create_sheet("README")
    for i, line in enumerate(README_TEXT.splitlines(), start=1):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions["A"].width = 90
    readme["A1"].font = Font(bold=True, size=12)

    # ---- Sheet 2: Crosswalk_Table ----
    ws_t = wb.create_sheet("Crosswalk_Table")
    headers = ["Estimate Item", "Bid Tab Count", "Project EAN", "Raw Bid Tab Item Description"]
    ws_t.append(headers)
    _hdr(ws_t)
    ws_t.freeze_panes = "A2"
    for _, r in table.iterrows():
        ws_t.append([r["Estimate Item"], r["Bid Tab Count"], r["Project EAN"],
                     r["Raw Bid Tab Item Description"]])
    ws_t.column_dimensions["A"].width = 45
    ws_t.column_dimensions["B"].width = 14
    ws_t.column_dimensions["C"].width = 55
    ws_t.column_dimensions["D"].width = 60

    # ---- Sheets 3–6 ----
    _df_sheet(wb, "Detail", detail)
    _df_sheet(wb, "Unmapped", unmapped)
    _df_sheet(wb, "Out_of_Catalog_Scope", oos)
    _df_sheet(wb, "Slicer_QA", slicer_qa)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


# ================================================================
# MAIN
# ================================================================

def main() -> None:
    global CSV_PATH, CATALOG_XLSX, OUTPUT_PATH

    parser = argparse.ArgumentParser(description="Build estimate item crosswalk artifacts.")
    parser.add_argument("--input_csv", default=str(CSV_PATH))
    parser.add_argument("--catalog_xlsx", default=str(CATALOG_XLSX))
    parser.add_argument("--out_dir", default=str(OUTPUT_PATH.parent))
    parser.add_argument("--out_prefix", default=OUTPUT_PATH.stem)
    args = parser.parse_args()

    CSV_PATH = Path(args.input_csv)
    CATALOG_XLSX = Path(args.catalog_xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH = out_dir / f"{args.out_prefix}.xlsx"

    print("Loading catalog and bid tabs ...")
    detail, table = run_crosswalk()

    n_total = len(detail)
    n_mapped = int((detail["mapping_status"] == "mapped").sum())
    n_unmapped = int((detail["mapping_status"] == "unmapped").sum())
    n_oos = int((detail["mapping_status"] == "out_of_catalog_scope").sum())

    print("\nWriting Excel workbook ...")
    write_xlsx(detail, table)
    detail.to_csv(out_dir / f"{args.out_prefix}_detail.csv", index=False)
    table.to_csv(out_dir / f"{args.out_prefix}_table.csv", index=False)
    detail[detail["mapping_status"] == "unmapped"].to_csv(out_dir / f"{args.out_prefix}_unmapped.csv", index=False)
    detail[detail["mapping_status"] == "out_of_catalog_scope"].to_csv(
        out_dir / f"{args.out_prefix}_out_of_catalog_scope.csv",
        index=False,
    )

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total bid tab rows (deduplicated): {n_total:,}")
    print(f"  Mapped:              {n_mapped:,}  ({100*n_mapped/n_total:.1f}%)")
    print(f"  Unmapped:            {n_unmapped:,}  ({100*n_unmapped/n_total:.1f}%)")
    print(f"  Out of catalog scope:{n_oos:,}  ({100*n_oos/n_total:.1f}%)")
    print(f"Crosswalk_Table rows:  {len(table):,}")
    print(f"Output file:           {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
