from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List, Sequence

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from .io_excel import load_workbook_data
from .normalize import canonicalize_name, clean_whitespace, load_name_map
from .parse_header import parse_project_header
from .parse_items import _extract_line_no, iter_item_rows
from .parse_table import detect_schedule_fields, parse_table_structure
from .qa import ParseIssue, enforce_canonical_key_uniqueness, write_parse_summary, write_qa_reports
from .schemas import SCHEMA_VERSION, coerce_compiled_record, compiled_columns

RAW_SNAPSHOT_COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row_index",
    "source_table_header_row",
    "project_ean",
    "solicitation_no",
    "letting_date_raw",
    "location_name_raw",
    "project_name_raw",
    "bid_schedule_name",
    "bid_schedule_type",
    "bid_schedule_code",
    "line_no",
    "item_description_raw",
    "quantity_raw",
    "unit_code_raw",
    "unit_price_raw",
    "total_price_raw",
    "bidder_name_raw",
    "is_totals_row",
    "totals_row_label",
    "schedule_total_raw",
]


def classify_bidder(name: str) -> tuple[str, bool]:
    lower = clean_whitespace(name).lower()
    is_est = "engineer" in lower and "estimate" in lower
    bidder_type = "ENGINEERS_ESTIMATE" if is_est else "CONTRACTOR"
    return bidder_type, is_est


def detect_candidate_sheet(ws: Worksheet) -> bool:
    item_rows = 0
    for r in range(1, ws.max_row + 1):
        row_vals = [clean_whitespace(ws.cell(row=r, column=c).value).lower() for c in range(1, ws.max_column + 1)]
        if "item no." in row_vals or "item no" in row_vals:
            unit_count = row_vals.count("unit price")
            total_count = row_vals.count("total price")
            if unit_count >= 2 and total_count >= 2:
                item_rows += 1
    return item_rows >= 1


def iter_workbooks(input_dir: Path | str) -> Iterable[Path]:
    root = Path(input_dir)
    for path in sorted(root.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        yield path


def build_raw_snapshot(workbooks: Sequence[Path] | Path | str, config: dict | None = None) -> pd.DataFrame:
    records: List[dict] = []
    paths = list(iter_workbooks(workbooks)) if isinstance(workbooks, (str, Path)) else list(workbooks)

    for path in paths:
        wb = load_workbook_data(path)
        for ws in wb.worksheets:
            if not detect_candidate_sheet(ws):
                continue

            structure = parse_table_structure(ws)
            if structure.table_header_row < 0 or not structure.bidder_blocks:
                continue

            header = parse_project_header(ws, structure.table_header_row)
            schedule = detect_schedule_fields(ws.title)

            row = structure.table_header_row + 1
            blank_streak = 0
            while row <= ws.max_row:
                line_no = _extract_line_no(ws, row, structure.col_map["line_no_col"])
                desc_raw = clean_whitespace(ws.cell(row=row, column=structure.col_map["desc_col"]).value)
                qty_raw = clean_whitespace(ws.cell(row=row, column=structure.col_map["qty_col"]).value)
                unit_raw = clean_whitespace(ws.cell(row=row, column=structure.col_map["unit_col"]).value)

                if not line_no and not desc_raw:
                    blank_streak += 1
                    if blank_streak >= 10:
                        break
                    row += 1
                    continue

                blank_streak = 0
                is_totals = any(token in desc_raw.lower() for token in ["total amount", "basis of award"])
                if not (line_no or is_totals):
                    row += 1
                    continue

                for block in structure.bidder_blocks:
                    unit_raw_cell = clean_whitespace(ws.cell(row=row, column=block.unit_price_col).value)
                    total_raw_cell = clean_whitespace(ws.cell(row=row, column=block.total_price_col).value)
                    records.append(
                        {
                            "source_file": path.name,
                            "source_sheet": ws.title,
                            "source_row_index": row,
                            "source_table_header_row": structure.table_header_row,
                            "project_ean": header.get("project_ean", ""),
                            "solicitation_no": header.get("solicitation_no", ""),
                            "letting_date_raw": header.get("letting_date_raw", ""),
                            "location_name_raw": header.get("location_name_raw", ""),
                            "project_name_raw": header.get("project_name_raw", ""),
                            "bid_schedule_name": schedule.get("bid_schedule_name", ""),
                            "bid_schedule_type": schedule.get("bid_schedule_type", ""),
                            "bid_schedule_code": schedule.get("bid_schedule_code", ""),
                            "line_no": "" if is_totals else line_no,
                            "item_description_raw": desc_raw,
                            "quantity_raw": "" if is_totals else qty_raw,
                            "unit_code_raw": "" if is_totals else unit_raw,
                            "unit_price_raw": "" if is_totals else unit_raw_cell,
                            "total_price_raw": "" if is_totals else total_raw_cell,
                            "bidder_name_raw": clean_whitespace(block.bidder_name_raw),
                            "is_totals_row": bool(is_totals),
                            "totals_row_label": desc_raw if is_totals else "",
                            "schedule_total_raw": total_raw_cell if is_totals else "",
                        }
                    )
                row += 1

    df = pd.DataFrame(records)
    if df.empty:
        return pd.DataFrame(columns=RAW_SNAPSHOT_COLUMNS)

    for c in RAW_SNAPSHOT_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df[RAW_SNAPSHOT_COLUMNS].sort_values(["source_file", "source_sheet", "source_row_index", "bidder_name_raw"], kind="stable").reset_index(drop=True)


def extract_file(path: Path, run_id: str, name_map: Dict[str, str], parse_issues: List[ParseIssue]) -> List[dict]:
    records: List[dict] = []
    wb = load_workbook_data(path)
    for ws in wb.worksheets:
        if not detect_candidate_sheet(ws):
            continue

        structure = parse_table_structure(ws)
        if structure.table_header_row < 0:
            parse_issues.append(ParseIssue(path.name, ws.title, "header_detection", ";".join(structure.warnings)))
            continue
        if structure.warnings:
            parse_issues.append(ParseIssue(path.name, ws.title, "parse_warning", ";".join(structure.warnings)))
        if not structure.bidder_blocks:
            continue

        header = parse_project_header(ws, structure.table_header_row)
        schedule = detect_schedule_fields(ws.title)

        for item in iter_item_rows(ws, structure.table_header_row, structure.col_map, structure.bidder_blocks):
            bidder_raw = item["bidder_name_raw"]
            bidder_canon = canonicalize_name(bidder_raw, name_map)
            bidder_type, is_est = classify_bidder(bidder_raw)
            parse_confidence = 0.95
            if bidder_raw == "UNKNOWN_BIDDER":
                parse_confidence -= 0.4
            if item["is_totals_row"]:
                parse_confidence -= 0.1

            record = {
                "schema_version": SCHEMA_VERSION,
                "extract_run_id": run_id,
                "source_file": path.name,
                "source_sheet": ws.title,
                "source_row_index": item["source_row_index"],
                "source_table_header_row": structure.table_header_row,
                **header,
                **schedule,
                "line_no": item.get("line_no", ""),
                "item_description_raw": item.get("item_description_raw", ""),
                "item_description_clean": item.get("item_description_clean", ""),
                "item_code_raw": item.get("item_code_raw", ""),
                "item_code_norm": item.get("item_code_norm", ""),
                "section_code_raw": item.get("section_code_raw", ""),
                "section_code_norm": item.get("section_code_norm", ""),
                "quantity": item.get("quantity"),
                "unit_code_raw": item.get("unit_code_raw", ""),
                "unit_code_norm": item.get("unit_code_norm", ""),
                "bidder_name_raw": bidder_raw,
                "bidder_name_canonical": bidder_canon,
                "bidder_type": bidder_type,
                "is_engineers_estimate": is_est,
                "unit_price": item.get("unit_price"),
                "total_price": item.get("total_price"),
                "total_price_calc": None,
                "price_valid_flag": False if is_est else True,
                "parse_confidence": max(0.0, min(1.0, parse_confidence)),
                "is_totals_row": item.get("is_totals_row", False),
                "totals_row_label": item.get("totals_row_label", ""),
                "schedule_total": item.get("schedule_total"),
                "specification": item.get("specification", ""),
                "alternate_specification": item.get("alternate_specification", ""),
                "pay_item_description": item.get("pay_item_description", ""),
                "supplemental_description": item.get("supplemental_description", ""),
                "item": item.get("item", ""),
            }
            records.append(coerce_compiled_record(record))
    return records


def build_clean_extract(raw_workbooks_or_snapshot: Sequence[Path] | Path | str, config_tables: dict | str | Path, run_id: str = "local", arith_tolerance: float = 0.02) -> tuple[pd.DataFrame, List[ParseIssue]]:
    config_dir = config_tables if isinstance(config_tables, (str, Path)) else "./config"
    name_map = load_name_map(str(config_dir)) if isinstance(config_dir, (str, Path)) else {}
    parse_issues: List[ParseIssue] = []
    all_records: List[dict] = []
    paths = list(iter_workbooks(raw_workbooks_or_snapshot)) if isinstance(raw_workbooks_or_snapshot, (str, Path)) else list(raw_workbooks_or_snapshot)
    for path in paths:
        all_records.extend(extract_file(Path(path), run_id, name_map, parse_issues))

    df = pd.DataFrame(all_records)
    if df.empty:
        return pd.DataFrame(columns=compiled_columns()), parse_issues

    for col in compiled_columns():
        if col not in df.columns:
            df[col] = None
    df = df[compiled_columns()]

    line_mask = df["is_totals_row"] == False  # noqa: E712
    for col in ["quantity", "unit_price", "total_price"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.loc[line_mask, "total_price_calc"] = df.loc[line_mask, "unit_price"] * df.loc[line_mask, "quantity"]

    delta = (df["total_price"] - df["total_price_calc"]).abs()
    threshold = df["total_price"].abs().fillna(0).clip(lower=1) * float(arith_tolerance)
    calc_valid = (delta <= threshold) | df["total_price"].isna() | df["unit_price"].isna() | df["quantity"].isna()
    df.loc[line_mask & (~df["is_engineers_estimate"].astype(bool)), "price_valid_flag"] = calc_valid
    df.loc[df["is_engineers_estimate"].astype(bool), "price_valid_flag"] = False

    str_cols = [c for c in compiled_columns() if df[c].dtype == "object"]
    for c in str_cols:
        df[c] = df[c].fillna("").astype(str)
        df[c] = df[c].replace({"nan": "", "None": "", "NA": "", "null": ""})

    return df, parse_issues


def write_extract_outputs(input_dir: Path | str, output_csv: Path | str, run_id: str, reports_dir: Path | str = "./reports", config_dir: Path | str = "./config", arith_tolerance: float = 0.02, totals_tolerance_amount: float = 5.0, totals_tolerance_pct: float = 0.02) -> pd.DataFrame:
    df, parse_issues = build_clean_extract(input_dir, config_dir, run_id=run_id, arith_tolerance=arith_tolerance)
    output_csv = Path(output_csv)
    reports_dir = Path(reports_dir)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_csv, index=False, na_rep="")

    write_parse_summary(parse_issues, reports_dir / "qa_parse_summary.csv")
    if not df.empty:
        write_qa_reports(df, reports_dir, tolerance=float(arith_tolerance), tolerance_amount=float(totals_tolerance_amount), tolerance_pct=float(totals_tolerance_pct))
        dupes = enforce_canonical_key_uniqueness(df)
        if not dupes.empty:
            parse_issues.append(ParseIssue("*", "*", "canonical_key_duplicates", f"rows={len(dupes)}"))
            write_parse_summary(parse_issues, reports_dir / "qa_parse_summary.csv")
    return df
