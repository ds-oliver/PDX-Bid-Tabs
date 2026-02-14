from __future__ import annotations

import re
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from .normalize import clean_whitespace

EAN_RE = re.compile(r"EAN\s+([A-Za-z0-9\-]+)", re.IGNORECASE)
SOL_RE = re.compile(r"Solicitation\s+No\.\s*[:\-]?\s*(.+)", re.IGNORECASE)
DATE_RE = re.compile(
    r"(\d{1,2}[\-/]\d{1,2}[\-/]\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},\s*\d{4}|\d{4}[\-/]\d{1,2}[\-/]\d{1,2})"
)


def _parse_date(value: str) -> Optional[str]:
    if not value:
        return None
    candidates = ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%b %d, %Y", "%B %d, %Y"]
    for fmt in candidates:
        try:
            return datetime.strptime(value.strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _normalize_location_text(value: str) -> str:
    text = clean_whitespace(value).upper()
    # Keep alphanumeric only so "HILLSBORO AIRPORT" -> "HILLSBOROAIRPORT".
    return re.sub(r"[^A-Z0-9]", "", text)


def parse_project_header(ws: Worksheet, table_header_row: int) -> Dict[str, str]:
    header_lines: List[str] = []
    max_scan_col = min(ws.max_column, 20)
    for r in range(1, table_header_row):
        vals = []
        for c in range(1, max_scan_col + 1):
            v = clean_whitespace(ws.cell(row=r, column=c).value)
            if v:
                vals.append(v)
        if vals:
            header_lines.append(" | ".join(vals))

    joined = "\n".join(header_lines)
    ean = ""
    sol = ""
    letting_raw = ""
    letting_date = ""

    ean_m = EAN_RE.search(joined)
    if ean_m:
        ean = clean_whitespace(ean_m.group(1))

    sol_m = SOL_RE.search(joined)
    if sol_m:
        sol = clean_whitespace(sol_m.group(1))

    date_m = DATE_RE.search(joined)
    if date_m:
        letting_raw = clean_whitespace(date_m.group(1))
        letting_date = _parse_date(letting_raw) or ""

    # Read left header block only to avoid contractor/bidder names from right-side columns.
    left_lines = []
    for r in range(1, table_header_row):
        v = clean_whitespace(ws.cell(row=r, column=1).value)
        if v and "item no" not in v.lower():
            left_lines.append(v)

    location = _normalize_location_text(left_lines[0]) if left_lines else ""

    project_lines = []
    for line in left_lines[1:]:
        l = line.lower()
        if l.startswith("ean ") or l.startswith("solicitation"):
            break
        if DATE_RE.search(line):
            break
        project_lines.append(line)

    project = clean_whitespace(" ".join(project_lines))

    return {
        "project_ean": ean,
        "solicitation_no": sol,
        "letting_date_raw": letting_raw,
        "letting_date": letting_date,
        "location_name_raw": location,
        "project_name_raw": project,
    }
