from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


MergedLookup = Dict[Tuple[int, int], Tuple[int, int, int, int]]


def load_workbook_data(path: Path):
    return load_workbook(path, data_only=True)


def build_merged_lookup(ws: Worksheet) -> MergedLookup:
    lookup: MergedLookup = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = (min_row, min_col, max_row, max_col)
    return lookup


def get_cell_value(ws: Worksheet, row: int, col: int):
    return ws.cell(row=row, column=col).value


def get_merged_top_left_value(ws: Worksheet, lookup: MergedLookup, row: int, col: int):
    bounds = lookup.get((row, col))
    if not bounds:
        return get_cell_value(ws, row, col)
    min_row, min_col, _, _ = bounds
    return get_cell_value(ws, min_row, min_col)


def iter_nonempty_rows(ws: Worksheet, min_row: int = 1, max_row: Optional[int] = None) -> Iterable[int]:
    max_r = max_row or ws.max_row
    for row in range(min_row, max_r + 1):
        row_vals = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in row_vals):
            yield row
