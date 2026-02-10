from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from .io_excel import MergedLookup, build_merged_lookup, get_merged_top_left_value
from .normalize import clean_whitespace


@dataclass
class BidderBlock:
    bidder_name_raw: str
    unit_price_col: int
    total_price_col: int


@dataclass
class TableParseResult:
    table_header_row: int
    col_map: Dict[str, int]
    bidder_blocks: List[BidderBlock]
    warnings: List[str]


HEADER_REQUIRED = {
    "item_no": "item no.",
    "item_desc": "item description",
    "qty": "estimated quantity",
    "units": "units",
}


def _norm(text: object) -> str:
    return clean_whitespace(text).lower()


def _detect_header_candidates(ws: Worksheet) -> List[int]:
    candidates = []
    for r in range(1, ws.max_row + 1):
        row_vals = [_norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if any(v == "item no." or v == "item no" for v in row_vals):
            unit_count = sum(1 for v in row_vals if v == "unit price")
            total_count = sum(1 for v in row_vals if v == "total price")
            if unit_count >= 2 and total_count >= 2:
                candidates.append(r)
    return candidates


def detect_table_header_row(ws: Worksheet) -> Tuple[Optional[int], str]:
    candidates = _detect_header_candidates(ws)
    if not candidates:
        return None, "missing_item_no"

    scored = []
    for r in candidates:
        row_vals = [_norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        score = sum(1 for v in HEADER_REQUIRED.values() if v in row_vals)
        scored.append((score, r))

    scored.sort(reverse=True)
    best_score, best_row = scored[0]
    tied = [row for score, row in scored if score == best_score]
    if len(tied) > 1:
        return None, "ambiguous_header_rows"
    return best_row, "ok"


def map_base_columns(ws: Worksheet, table_header_row: int) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row=table_header_row, column=c).value)
        if v in {"item no.", "item no"}:
            col_map["line_no_col"] = c
        elif v == "item description":
            col_map["desc_col"] = c
        elif v == "estimated quantity":
            col_map["qty_col"] = c
        elif v == "units":
            col_map["unit_col"] = c
    return col_map


def _find_header_name_for_pair(
    ws: Worksheet,
    merged_lookup: MergedLookup,
    table_header_row: int,
    unit_col: int,
    total_col: int,
    lookback_rows: int,
) -> str:
    for r in range(table_header_row - 1, max(0, table_header_row - lookback_rows) - 1, -1):
        v_unit = clean_whitespace(get_merged_top_left_value(ws, merged_lookup, r, unit_col))
        v_total = clean_whitespace(get_merged_top_left_value(ws, merged_lookup, r, total_col))

        rng_u = merged_lookup.get((r, unit_col))
        rng_t = merged_lookup.get((r, total_col))
        if rng_u and rng_t and rng_u == rng_t:
            min_row, min_col, max_row, max_col = rng_u
            if min_col <= unit_col and max_col >= total_col:
                return clean_whitespace(ws.cell(row=min_row, column=min_col).value)

        if v_unit and v_total and v_unit == v_total:
            return v_unit
        if v_unit and not v_total:
            return v_unit
        if v_total and not v_unit:
            return v_total
    return "UNKNOWN_BIDDER"


def detect_bidder_blocks(ws: Worksheet, table_header_row: int, lookback_rows: int = 10) -> Tuple[List[BidderBlock], List[str]]:
    merged_lookup = build_merged_lookup(ws)
    unit_cols = []
    total_cols = []
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row=table_header_row, column=c).value)
        if v == "unit price":
            unit_cols.append(c)
        elif v == "total price":
            total_cols.append(c)

    warnings: List[str] = []
    if len(unit_cols) != len(total_cols):
        warnings.append("bidder_pair_count_mismatch")

    n = min(len(unit_cols), len(total_cols))
    blocks: List[BidderBlock] = []
    for i in range(n):
        u = unit_cols[i]
        t = total_cols[i]
        if t < u:
            warnings.append(f"misordered_pair_{u}_{t}")
            continue
        bidder_name = _find_header_name_for_pair(ws, merged_lookup, table_header_row, u, t, lookback_rows)
        blocks.append(BidderBlock(bidder_name_raw=bidder_name, unit_price_col=u, total_price_col=t))
    return blocks, warnings


def detect_schedule_fields(sheet_name: str) -> Dict[str, str]:
    name = clean_whitespace(sheet_name)
    lower = name.lower()
    schedule_type = "ALTERNATE" if "alt" in lower or "alternate" in lower else "BASE"

    code = ""
    m = re.search(r"(?:alt(?:ernate)?)\s*[-_ ]*([A-Za-z0-9]+)", lower)
    if m:
        code = f"ALT_{m.group(1).upper()}"

    return {
        "bid_schedule_name": name,
        "bid_schedule_type": schedule_type,
        "bid_schedule_code": code,
    }


def parse_table_structure(ws: Worksheet) -> TableParseResult:
    warnings: List[str] = []
    row, status = detect_table_header_row(ws)
    if row is None:
        return TableParseResult(table_header_row=-1, col_map={}, bidder_blocks=[], warnings=[status])

    col_map = map_base_columns(ws, row)
    missing = [k for k in ["line_no_col", "desc_col", "qty_col", "unit_col"] if k not in col_map]
    if missing:
        warnings.append("missing_base_columns:" + ",".join(missing))

    bidder_blocks, bidder_warnings = detect_bidder_blocks(ws, row)
    warnings.extend(bidder_warnings)
    if not bidder_blocks:
        warnings.append("no_bidder_blocks")

    return TableParseResult(table_header_row=row, col_map=col_map, bidder_blocks=bidder_blocks, warnings=warnings)
