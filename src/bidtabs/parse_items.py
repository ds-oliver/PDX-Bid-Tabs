from __future__ import annotations

import csv
import re
from decimal import Decimal
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from .normalize import clean_whitespace, normalize_unit_code, parse_decimal
from .parse_table import BidderBlock

TOTALS_PATTERN = re.compile(r"(total amount|basis of award)", re.IGNORECASE)
PAREN_PATTERN = re.compile(r"\([^)]*\)")
SECTION_WORD_PATTERN = re.compile(r"sections?", re.IGNORECASE)
ITEM_WORD_PATTERN = re.compile(r"item", re.IGNORECASE)
SECTION_CODE_PATTERN = re.compile(r"\d{6}")
ITEM_TOKEN_CODE_PATTERN = re.compile(r"item\s*([A-Z]{1,2}\s*-\s*\d{3})", re.IGNORECASE)
TRAILING_QUALIFIER_PATTERN = re.compile(r"^(.*)\(([^()]*)\)\s*$")
LIST_GUARD_PATTERN = re.compile(r",[^,]*(?:\band\b|&)[^,]*$", re.IGNORECASE)

CAMELCASE_BOUNDARY_PATTERN = re.compile(r"(?<=[a-z])(?=[A-Z])")
SPEC_TOKEN_SPACE_PATTERN = re.compile(r"(?i)\b(Section|Sections|Item)(\d)")
ADJUSTMENT_TO_GRADE_PATTERN = re.compile(r"(?i)AdjustmenttoGrade")
SEPARATOR_DASH_NORMALIZE_PATTERN = re.compile(r"\b([A-Za-z]{3,})-(?=[A-Z0-9])")

LEADING_MEASUREMENT_PATTERN = re.compile(
    r"^\s*("
    r"\d+(?:\.\d+)?(?:\s*-\s*\d+(?:\.\d+)?)?(?:/\d+(?:\.\d+)?)?"
    r"\s*(?:-|\s)\s*"
    r"(?:INCH(?:ES)?|IN\.?|\"|FOOT|FEET|FT|SF|SY|LF|CY|EA|EACH|LB|TON|GAL(?:LON)?S?)"
    r")\b\s*",
    re.IGNORECASE,
)

MEASUREMENT_PATTERN = re.compile(
    r"(" \
    r"\b\d+(?:\.\d+)?\s*-\s*\d+(?:\.\d+)?\s*(?:SF|SY|LF|CY|EA|EACH|LB|TON|GALLON|GALLONS|GAL|INCH|INCHES|IN\.?|\"|FT|FEET|FOOT|AMP|AMPS|VOLT|VOLTS|WATT|WATTS|HP|KVA)\b" \
    r"|\b\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)?\s*[-\s]*\s*(?:\"|INCH(?:ES)?|IN\.?|FT|FEET|FOOT|SF|SY|LF|CY|EA|EACH|LB|TON|GAL(?:LON)?S?|AMP(?:S)?|VOLT(?:S)?|WATT(?:S)?|HP|KVA)\b" \
    r"|\b\d+\s*-\s*\d+\b" \
    r"|\bDEPTH\b" \
    r")",
    re.IGNORECASE,
)

VARIANT_KEYWORDS = {
    "ORIGINAL",
    "RENEWAL",
    "SILICONE",
    "CLEANOUT",
    "MINOR ADJUSTMENT",
    "MAJOR ADJUSTMENT",
}

SPEC_LIKE_CODE_PATTERN = re.compile(r"^[A-Z]{1,2}-\d{2,6}$")
BASE_PLUS_GLUED_QUAL_PATTERN = re.compile(r"^(?P<base>[A-Za-z]{2,})(?P<qual>\d.+)$")
GLUED_QUAL_PLUS_BASE_PATTERN = re.compile(r"^(?P<qual>\d(?:[^A-Za-z]|[A-Z0-9'\"/])*)(?P<base>[A-Z][a-z].*)$")
COMPACT_DESCRIPTOR_PATTERN = re.compile(r"^[A-Za-z0-9'\"/,-]+$")

OVERRIDE_FILE = Path(__file__).resolve().parents[2] / "config" / "pay_item_parse_overrides.csv"


@lru_cache(maxsize=1)
def _load_parse_overrides() -> List[Dict[str, str]]:
    if not OVERRIDE_FILE.exists():
        return []

    out: List[Dict[str, str]] = []
    with OVERRIDE_FILE.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pattern = clean_whitespace(row.get("raw_description_pattern", ""))
            if not pattern:
                continue
            out.append(
                {
                    "raw_description_pattern": pattern,
                    "force_pay_item_description": clean_whitespace(row.get("force_pay_item_description", "")),
                    "force_supplemental_description": clean_whitespace(row.get("force_supplemental_description", "")),
                }
            )
    return out


def normalize_description_for_parsing(raw_text: str) -> str:
    text = clean_whitespace(raw_text)
    if not text:
        return ""

    text = SPEC_TOKEN_SPACE_PATTERN.sub(r"\1 \2", text)
    text = ADJUSTMENT_TO_GRADE_PATTERN.sub("Adjustment to Grade", text)
    text = SEPARATOR_DASH_NORMALIZE_PATTERN.sub(r"\1 - ", text)

    return clean_whitespace(text)

def _decamel_if_needed(s: str) -> str:
    value = clean_whitespace(s)
    if not value:
        return value

    parts = []
    for token in value.split(" "):
        if sum(1 for _ in CAMELCASE_BOUNDARY_PATTERN.finditer(token)) >= 2:
            parts.append(CAMELCASE_BOUNDARY_PATTERN.sub(" ", token))
        else:
            parts.append(token)
    return clean_whitespace(" ".join(parts))


def _is_compact_qualifier_token(s: str) -> bool:
    value = clean_whitespace(s)
    if not value:
        return False
    if " " in value:
        return False
    if SPEC_LIKE_CODE_PATTERN.match(value.upper()):
        return False
    if MEASUREMENT_PATTERN.search(value):
        return True
    if not COMPACT_DESCRIPTOR_PATTERN.match(value):
        return False
    if sum(1 for _ in CAMELCASE_BOUNDARY_PATTERN.finditer(value)) >= 1:
        return True
    if any(ch.isdigit() for ch in value):
        return True
    return False


def _rhs_is_qualifier(rhs: str) -> bool:
    value = clean_whitespace(rhs)
    if not value:
        return False

    upper = value.upper()
    if upper in VARIANT_KEYWORDS:
        return True

    if MEASUREMENT_PATTERN.search(value):
        return True

    return False


def _extract_inline_paren_qualifiers(core: str) -> Tuple[str, List[str]]:
    supplements: List[str] = []
    if not core:
        return core, supplements

    parts: List[str] = []
    idx = 0
    for m in PAREN_PATTERN.finditer(core):
        parts.append(core[idx : m.start()])
        inner = core[m.start() + 1 : m.end() - 1]
        inner_clean = clean_whitespace(inner)

        if SECTION_WORD_PATTERN.search(inner_clean) or ITEM_WORD_PATTERN.search(inner_clean):
            parts.append(core[m.start() : m.end()])
        else:
            if _rhs_is_qualifier(inner_clean) or _is_compact_qualifier_token(inner_clean):
                supplements.append(inner_clean)
            else:
                parts.append(core[m.start() : m.end()])

        idx = m.end()

    parts.append(core[idx:])
    return clean_whitespace("".join(parts)), supplements


def _split_item_and_supplemental(desc_core: str) -> Tuple[str, str]:
    core = clean_whitespace(desc_core)
    if not core:
        return "", ""

    supplemental_parts: List[str] = []

    # Leading measurement token (e.g., 10-Foot ...)
    m = LEADING_MEASUREMENT_PATTERN.match(core)
    if m:
        token = clean_whitespace(m.group(1))
        token = re.sub(r"\s*-\s*", "-", token)
        supplemental_parts.append(token)
        core = clean_whitespace(core[m.end() :])

    # InstallPF12"Diameter,... -> base + glued qualifier
    m = BASE_PLUS_GLUED_QUAL_PATTERN.match(core)
    if m:
        base = clean_whitespace(m.group("base"))
        qual = clean_whitespace(m.group("qual"))
        if _is_compact_qualifier_token(qual) or "," in qual:
            supplemental_parts.append(qual)
            return base, " | ".join([p for p in supplemental_parts if p])

    # 8'HX14'WCulvert / 3'HighBlue... -> qualifier prefix + base
    m = GLUED_QUAL_PLUS_BASE_PATTERN.match(core)
    if m:
        qual = clean_whitespace(m.group("qual"))
        base = clean_whitespace(m.group("base"))
        if _is_compact_qualifier_token(qual) or MEASUREMENT_PATTERN.search(qual):
            supplemental_parts.append(qual)
            return base, " | ".join([p for p in supplemental_parts if p])

    # Inline non-spec parenthetical qualifiers
    core, paren_supps = _extract_inline_paren_qualifiers(core)
    supplemental_parts.extend(paren_supps)

    # Dash separator qualifier
    if " - " in core:
        lhs, rhs = core.rsplit(" - ", 1)
        lhs = clean_whitespace(lhs)
        rhs = clean_whitespace(rhs)
        if _rhs_is_qualifier(rhs) or _is_compact_qualifier_token(rhs):
            supplemental_parts.append(rhs)
            return _decamel_if_needed(lhs), " | ".join([p for p in supplemental_parts if p])
    # Comma logic with list guard
    if "," in core:
        comma_count = core.count(",")
        if comma_count >= 2 and LIST_GUARD_PATTERN.search(core):
            return _decamel_if_needed(core), " | ".join([p for p in supplemental_parts if p])

        lhs, rhs = core.rsplit(",", 1)
        lhs = clean_whitespace(lhs)
        rhs = clean_whitespace(rhs)
        if _rhs_is_qualifier(rhs) or _is_compact_qualifier_token(rhs):
            supplemental_parts.append(rhs)
            return _decamel_if_needed(lhs), " | ".join([p for p in supplemental_parts if p])

    if not supplemental_parts:
        m = TRAILING_QUALIFIER_PATTERN.match(core)
        if m:
            base = clean_whitespace(m.group(1))
            qualifier = clean_whitespace(m.group(2))
            if base and qualifier and (_rhs_is_qualifier(qualifier) or _is_compact_qualifier_token(qualifier)):
                supplemental_parts.append(qualifier)
                return _decamel_if_needed(base), " | ".join([p for p in supplemental_parts if p])

    return _decamel_if_needed(core), " | ".join([p for p in supplemental_parts if p])


def _apply_parse_overrides(raw_text: str, parsed: Dict[str, str]) -> Dict[str, str]:
    text = clean_whitespace(raw_text)
    for rule in _load_parse_overrides():
        pattern = rule["raw_description_pattern"]
        try:
            matched = re.search(pattern, text, flags=re.IGNORECASE) is not None
        except re.error:
            matched = pattern.lower() in text.lower()
        if not matched:
            continue

        forced_item = rule["force_pay_item_description"]
        forced_supp = rule["force_supplemental_description"]
        if forced_item:
            parsed["item"] = forced_item
        if forced_supp or forced_supp == "":
            parsed["supplemental_description"] = forced_supp

        spec = parsed.get("spec_code_primary", "UNCLASSIFIED")
        base = parsed.get("item", "")
        parsed["item_display"] = f"{spec} - {base}" if base else spec
        parsed["parse_type"] = f"{parsed.get('parse_type', 'UNCLASSIFIED')}_OVERRIDE"
        return parsed

    return parsed


def _parse_single_variant(raw_text: str):
    text = normalize_description_for_parsing(raw_text)
    parens = list(PAREN_PATTERN.finditer(text))

    if not parens:
        item_base, supplemental = _split_item_and_supplemental(text)
        return {
            "spec_code_primary": "UNCLASSIFIED",
            "spec_code_alternates": "",
            "desc_core": text,
            "item": item_base,
            "supplemental_description": supplemental,
            "item_display": f"UNCLASSIFIED - {item_base}" if item_base else "UNCLASSIFIED",
            "parse_type": "UNCLASSIFIED",
        }

    token_match = None
    for m in reversed(parens):
        token = m.group(0)
        if SECTION_WORD_PATTERN.search(token) or ITEM_WORD_PATTERN.search(token):
            trailing = clean_whitespace(text[m.end() :])
            if trailing == "":
                token_match = m
                break

    if token_match is None:
        item_base, supplemental = _split_item_and_supplemental(text)
        return {
            "spec_code_primary": "UNCLASSIFIED",
            "spec_code_alternates": "",
            "desc_core": text,
            "item": item_base,
            "supplemental_description": supplemental,
            "item_display": f"UNCLASSIFIED - {item_base}" if item_base else "UNCLASSIFIED",
            "parse_type": "UNCLASSIFIED",
        }

    token = token_match.group(0)
    desc_core = clean_whitespace(text[: token_match.start()])

    if SECTION_WORD_PATTERN.search(token):
        codes = SECTION_CODE_PATTERN.findall(token)
        ordered: List[str] = []
        for code in codes:
            if code not in ordered:
                ordered.append(code)
        if ordered:
            primary = ordered[0]
            alternates = " | ".join(ordered[1:])
            item_base, supplemental = _split_item_and_supplemental(desc_core)
            return {
                "spec_code_primary": primary,
                "spec_code_alternates": alternates,
                "desc_core": desc_core,
                "item": item_base,
                "supplemental_description": supplemental,
                "item_display": f"{primary} - {item_base}" if item_base else primary,
                "parse_type": "SECTION",
            }

    if ITEM_WORD_PATTERN.search(token):
        item_match = ITEM_TOKEN_CODE_PATTERN.search(token)
        if item_match:
            primary = item_match.group(1).upper()
            primary = re.sub(r"\s*-\s*", "-", primary)
            item_base, supplemental = _split_item_and_supplemental(desc_core)
            return {
                "spec_code_primary": primary,
                "spec_code_alternates": "",
                "desc_core": desc_core,
                "item": item_base,
                "supplemental_description": supplemental,
                "item_display": f"{primary} - {item_base}" if item_base else primary,
                "parse_type": "ITEM",
            }

    item_base, supplemental = _split_item_and_supplemental(text)
    return {
        "spec_code_primary": "UNCLASSIFIED",
        "spec_code_alternates": "",
        "desc_core": text,
        "item": item_base,
        "supplemental_description": supplemental,
        "item_display": f"UNCLASSIFIED - {item_base}" if item_base else "UNCLASSIFIED",
        "parse_type": "UNCLASSIFIED",
    }


def parse_item_fields(raw: str) -> Dict[str, str]:
    raw_norm = clean_whitespace(raw)
    candidates = [c.strip() for c in str(raw_norm).split("|")]
    candidates = [c for c in candidates if c] or [raw_norm]

    parsed = [_parse_single_variant(c) for c in candidates]
    chosen = parsed[0]
    for result in parsed:
        if result["parse_type"] != "UNCLASSIFIED":
            chosen = result
            break

    return _apply_parse_overrides(raw_norm, chosen)


def parse_description_components(raw: str):
    fields = parse_item_fields(raw)
    return (
        fields["spec_code_primary"],
        fields["spec_code_alternates"],
        fields["desc_core"],
        fields["parse_type"],
    )


def extract_item_tokens(description: str) -> Dict[str, str]:
    fields = parse_item_fields(description)
    spec = fields["spec_code_primary"]
    parse_type = fields["parse_type"]

    item_code_raw = ""
    section_code_raw = ""

    if parse_type.startswith("ITEM"):
        item_code_raw = clean_whitespace(spec)
    elif parse_type.startswith("SECTION"):
        section_code_raw = clean_whitespace(spec)

    item_code_norm = item_code_raw.upper().replace(" ", "") if item_code_raw else ""
    section_code_norm = re.sub(r"\D", "", section_code_raw) if section_code_raw else ""

    return {
        "item_code_raw": item_code_raw,
        "item_code_norm": item_code_norm,
        "section_code_raw": section_code_raw,
        "section_code_norm": section_code_norm,
    }


def _decimal_to_float(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def _extract_line_no(ws: Worksheet, row: int, col: int) -> str:
    cell = ws.cell(row=row, column=col)
    raw = cell.value
    if raw is None:
        return ""
    if isinstance(raw, str):
        return clean_whitespace(raw)
    if isinstance(raw, (int, float)):
        num_format = clean_whitespace(cell.number_format)
        if num_format and set(num_format) <= {"0"}:
            width = len(num_format)
            try:
                return str(int(raw)).zfill(width)
            except Exception:
                pass
        if int(raw) == raw:
            return str(int(raw))
        return clean_whitespace(raw)
    return clean_whitespace(raw)


def iter_item_rows(
    ws: Worksheet,
    table_header_row: int,
    col_map: Dict[str, int],
    bidder_blocks: List[BidderBlock],
    termination_blank_streak: int = 10,
) -> Iterable[dict]:
    blank_streak = 0
    row = table_header_row + 1

    while row <= ws.max_row:
        line_no_raw = ws.cell(row=row, column=col_map["line_no_col"]).value
        desc_raw = ws.cell(row=row, column=col_map["desc_col"]).value
        qty_raw = ws.cell(row=row, column=col_map["qty_col"]).value
        unit_raw = ws.cell(row=row, column=col_map["unit_col"]).value

        line_no = _extract_line_no(ws, row, col_map["line_no_col"])
        desc = clean_whitespace(desc_raw)

        if not line_no and not desc:
            blank_streak += 1
            if blank_streak >= termination_blank_streak:
                break
            row += 1
            continue

        blank_streak = 0
        is_totals = bool(TOTALS_PATTERN.search(desc))
        tokens = extract_item_tokens(desc)
        parsed_fields = parse_item_fields(desc)
        qty = parse_decimal(qty_raw)
        unit_norm = normalize_unit_code(unit_raw) if unit_raw is not None else ""

        if line_no or is_totals:
            for block in bidder_blocks:
                unit_price = parse_decimal(ws.cell(row=row, column=block.unit_price_col).value)
                total_price = parse_decimal(ws.cell(row=row, column=block.total_price_col).value)
                schedule_total = total_price if is_totals else None

                spec = "" if is_totals else parsed_fields.get("spec_code_primary", "")
                alt_spec = "" if is_totals else parsed_fields.get("spec_code_alternates", "")
                pay_item_desc = "" if is_totals else parsed_fields.get("item", "")
                supp_desc = "" if is_totals else parsed_fields.get("supplemental_description", "")
                item_value = "" if is_totals else (f"{spec} - {pay_item_desc}" if spec and pay_item_desc else pay_item_desc)

                record = {
                    "source_row_index": row,
                    "line_no": "" if is_totals else line_no,
                    "item_description_raw": desc,
                    "item_description_clean": desc if is_totals else parsed_fields["desc_core"],
                    "quantity": None if is_totals else _decimal_to_float(qty),
                    "unit_code_raw": "" if is_totals else clean_whitespace(unit_raw),
                    "unit_code_norm": "" if is_totals else unit_norm,
                    "unit_price": None if is_totals else _decimal_to_float(unit_price),
                    "total_price": None if is_totals else _decimal_to_float(total_price),
                    "schedule_total": _decimal_to_float(schedule_total),
                    "is_totals_row": is_totals,
                    "totals_row_label": desc if is_totals else "",
                    "bidder_name_raw": block.bidder_name_raw,
                    "specification": spec,
                    "alternate_specification": alt_spec,
                    "pay_item_description": pay_item_desc,
                    "supplemental_description": supp_desc,
                    "item": item_value,
                    **tokens,
                }
                yield record
        row += 1
